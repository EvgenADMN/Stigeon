from datetime import date, datetime
from django.http import HttpResponse
from pandas import date_range
from django.shortcuts import render, redirect
from django.apps import apps
import win32com.client as w32
import pythoncom
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
import os

Patterns = apps.get_model('patterns', 'Patterns')
Patterns_Devices = apps.get_model('patterns', 'Patterns_Devices')
Patterns_Employees = apps.get_model('patterns', 'Patterns_Employees')
Periods = apps.get_model('patterns', 'Periods')
Absenses = apps.get_model('patterns', 'Absenses')
PercoConnect = apps.get_model('settings', 'PercoConnect')
Employees = apps.get_model('settings', 'Employees')
Devices = apps.get_model('settings', 'Devices')
Colors = apps.get_model('settings', 'Colors')


class CheckupOnce(Exception):
    pass


def checkContent(data, params):
    print('------DATA:------')
    for k, v in data.items():
        print(f'{k}: {v}')

    print('------PARAMS:------')
    for k, v in params.items():
        print(f'{k}: {v}')


def fillParams(data, params):
    days_amount = (date.fromisoformat(params['date-to']) - date.fromisoformat(params['date-from'])).days + 1
    data['days'] = {}
    for day in date_range(params['date-from'], periods=days_amount):
        data['days'][str(date.fromisoformat(str(day).split()[0]))] = date.fromisoformat(str(day).split()[0])

    data['devices'] = [device.device_id for device in
                       Patterns_Devices.objects.filter(pattern_name=Patterns.objects.get(id=params['pattern']))]
    data['persons'] = {}
    for person in Patterns_Employees.objects.filter(pattern_name=Patterns.objects.get(id=params['pattern'])):
        data['persons'][Employees.objects.get(
            id=person.employee_id.id)] = f'{Employees.objects.get(id=person.employee_id.id).last_name}  ' \
                                         f'{Employees.objects.get(id=person.employee_id.id).first_name}  ' \
                                         f'{Employees.objects.get(id=person.employee_id.id).middle_name}'


def connectPerco(PercoConnect):
    pythoncom.CoInitialize()
    setting = PercoConnect.objects.all()[0]
    host = setting.host
    port = setting.port
    login = setting.login
    passwd = setting.password

    return {'dispatch': w32.Dispatch(setting.sdk_name),
            'host': host,
            'port': port,
            'login': login,
            'passwd': passwd,
            }


def getGraphFromPerco(com, id_internal, date_from, date_to):
    XML_DOM = w32.Dispatch('Microsoft.XMLDOM')
    Header = XML_DOM.createProcessingInstruction('xml', 'version="1.0" encoding="UTF-8" standalone="yes"')
    XML_DOM.appendChild(Header)

    Elem = XML_DOM.createElement("documentrequest")
    Elem.setAttribute("type", "regevents")
    NodeRoot = XML_DOM.appendChild(Elem)

    Elem = XML_DOM.createElement("eventsreport")
    Elem.setAttribute("beginperiod", f"{date_from[8:10]}.{date_from[5:7]}.{date_from[:4]}")
    Elem.setAttribute("endperiod", f"{date_to[8:10]}.{date_to[5:7]}.{date_to[:4]}")
    Elem.setAttribute("beginperiodtime", "00:00")
    Elem.setAttribute("endperiodtime", "00:00")
    Elem.setAttribute("id_device_internal", id_internal)  # f"{id_internal}")
    NodeRoot.appendChild(Elem)

    com['dispatch'].SetConnect(com['host'], com['port'], com['login'], com['passwd'])
    com['dispatch'].GetData(XML_DOM)
    XML_DOM.save('schedule/files/events.xml')


def parceXML(schedule):
    root = ET.parse('schedule/files/events.xml').getroot()

    for event in root.findall('eventsreport/events/event'):
        f_fio = event.attrib['f_fio']
        f_time_ev = event.attrib['f_time_ev']
        f_id_staff_internal = event.attrib['f_id_staff_internal']
        f_date_ev = event.attrib['f_date_ev']
        formated_date = f"{f_date_ev.split('.')[2]}-{f_date_ev.split('.')[1]}-{f_date_ev.split('.')[0]}"

        if all((f_fio, f_time_ev, f_id_staff_internal)):
            if f_id_staff_internal in [str(person.id_internal) for person in schedule]:
                for person in schedule:
                    if str(person.id_internal) == f_id_staff_internal and formated_date in schedule[person]:
                        schedule[person][formated_date].append(f_time_ev)
        else:
            pass

    for name in schedule:
        for day in schedule[name]:
            schedule[name][day].sort()


def col_letters(days, start_num=1, start_letter=None, day_letter = {}):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'  # 26
    for day in range(days):
        try:
            if start_letter is not None:
                day_letter[day + start_num] = letters[start_letter] + letters[day]
                continue
            day_letter[day + start_num] = letters[day]
        except IndexError:
            if start_letter is None:
                start_letter = 0
            else:
                start_letter += 1

            col_letters(days=days-day, start_num=(day + 1) * (start_letter + 1), start_letter=start_letter, day_letter=day_letter)
            break

    return day_letter


def fillExcel(data, params):
    weekdays = {
        0: 'Понедельник',
        1: 'Вторник',
        2: 'Среда',
        3: 'Четверг',
        4: 'Пятница',
        5: 'Суббота',
        6: 'Воскресение'
    }
    excel_file = f'schedule/files/schedule_{date.today().strftime("%d_%m_%Y")}.xlsx'
    workbook = openpyxl.Workbook()
    table = workbook.active
    table.title = 'График сотрудников'

    table.freeze_panes = "B3"


    column = 2
    for day in data['days']:
        table.merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + 1)
        table.cell(row=1, column=column).alignment = Alignment(horizontal='center')
        table.merge_cells(start_row=2, start_column=column, end_row=2, end_column=column + 1)
        table.cell(row=2, column=column).alignment = Alignment(horizontal='center')
        table.cell(row=1, column=column, value=day)
        table.cell(row=2, column=column, value=weekdays[data['days'][day].weekday()])
        table.cell(row=2, column=column).border = Border(bottom=Side(style='medium'))
        table.cell(row=2, column=column + 1).border = Border(bottom=Side(style='medium'))
        column += 2

    row = 3
    for person in data['persons']:
        table.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=1)
        table.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        table.cell(row=row, column=1).border = Border(right=Side(style='medium'))
        table.cell(row=row + 1, column=1).border = Border(right=Side(style='medium'))
        table.cell(row=row + 2, column=1).border = Border(right=Side(style='medium'))
        table.cell(row=row, column=1, value=f'{person.last_name} {person.first_name} {person.middle_name}')
        row += 3

    row = 3
    columns = col_letters(len(data['days']) * 2)
    for person in data['persons']:
        column = 2
        for day in data['days']:
            table.merge_cells(start_row=row + 2, start_column=column, end_row=row + 2, end_column=column + 1)
            table.cell(row=row + 2, column=column).alignment = Alignment(horizontal='center')
            table.cell(row=row + 2, column=column).border = Border(bottom=Side(style='thin'))
            table.cell(row=row + 2, column=column + 1).border = Border(bottom=Side(style='thin'))
            table.cell(row=row, column=column + 1).border = Border(right=Side(style='thin'))
            table.cell(row=row + 1, column=column + 1).border = Border(right=Side(style='thin'))
            table.cell(row=row + 2, column=column + 1).border = Border(right=Side(style='thin'))
            table.row_dimensions[row + 2].height = 45
            if table.column_dimensions[columns[list(data['days'].keys()).index(day) + 1]] != 1:
                table.column_dimensions[columns[list(data['days'].keys()).index(day) + 1]].width = 12
                table.column_dimensions[columns[list(data['days'].keys()).index(day) + 2]].width = 12

            for graph in data['graphs']:
                if graph.split('_')[2] == day and int(graph.split('_')[1]) == person.id:
                    if data['graphs'][graph] == 'dayoff':
                        table.cell(row=row + 2, column=column, value='Выходной')
                        continue
                    elif 'abs' in data['graphs'][graph]:
                        table.cell(row=row + 2, column=column, value=Absenses.objects.get(id=int(data['graphs'][graph].split('_')[1])).name)
                        table.cell(row=row + 2, column=column).fill = PatternFill('solid',
                                                                                  fgColor=Colors.objects.get().purpose[
                                                                                          1:])
                        continue
                    table.cell(row=row, column=column,
                               value=Periods.objects.get(id=int(data['graphs'][graph])).time_from)
                    table.cell(row=row, column=column + 1,
                               value=Periods.objects.get(id=int(data['graphs'][graph])).time_to)

            try:
                table.cell(row=row + 1, column=column, value=data['schedule'][person][day][0])
                if data['schedule'][person][day][0] == data['schedule'][person][day][-1]:
                    raise CheckupOnce
                table.cell(row=row + 1, column=column + 1, value=data['schedule'][person][day][-1])

                if table.cell(row=row + 2, column=column).value == 'Выходной' and table.cell(row=row + 1, column=column + 1).value:
                    table.cell(row=row + 2, column=column, value='Был в выходной!')

            except IndexError:
                if table.cell(row=row + 2, column=column).value is None:
                    table.cell(row=row + 2, column=column, value='Пропуск!')
                    table.cell(row=row + 2, column=column).fill = PatternFill('solid',
                                                                              fgColor=Colors.objects.get().absenteeism[
                                                                                      1:])

            except CheckupOnce:
                if table.cell(row=row + 2, column=column).value is None:
                    table.cell(row=row + 2, column=column, value='Одна отметка!')
                    table.cell(row=row + 2, column=column).fill = PatternFill('solid',
                                                                              fgColor=Colors.objects.get().absenteeism[
                                                                                      1:])
                elif table.cell(row=row + 2, column=column).value not in [absense.name for absense in Absenses.objects.all()]:
                    table.cell(row=row + 2, column=column, value='Был в выходной!')

            if table.cell(row=row + 2, column=column).value is None:
                late = ''
                early = ''

                sch_time_from_cell = table.cell(row=row, column=column).value
                sch_time_to_cell = table.cell(row=row, column=column + 1).value
                fact_time_from_cell = table.cell(row=row + 1, column=column).value
                fact_time_to_cell = table.cell(row=row + 1, column=column + 1).value

                sch_time_from = datetime.strptime(str(sch_time_from_cell), "%H:%M:%S")
                sch_time_to = datetime.strptime(str(sch_time_to_cell), "%H:%M:%S")
                fact_time_from = datetime.strptime(str(fact_time_from_cell), "%H:%M:%S")
                fact_time_to = datetime.strptime(str(fact_time_to_cell), "%H:%M:%S")

                if str(sch_time_from - fact_time_from)[0] == '-':
                    late = f'Опоздание: {fact_time_from - sch_time_from}\n'
                    table.cell(row=row + 2, column=column).fill = PatternFill('solid',
                                                                              fgColor=Colors.objects.get().lateness[
                                                                                      1:])
                if str(fact_time_to - sch_time_to)[0] == '-':
                    early = f'Уход: {sch_time_to - fact_time_to}\n'
                    table.cell(row=row + 2, column=column).fill = PatternFill('solid',
                                                                              fgColor=Colors.objects.get().early[
                                                                                      1:])

                work_time = f'Всего: {fact_time_to - fact_time_from}'
                if late and early:
                    table.cell(row=row + 2, column=column).fill = PatternFill('solid',
                                                                              fgColor=Colors.objects.get().lateness_early[
                                                                                      1:])

                table.cell(row=row + 2, column=column, value=f'{late}{early}{work_time}')

            column += 2
        row += 3

    table.column_dimensions['A'].width = 35
    workbook.save(excel_file)


# ================================


def index(request):
    data = {
        'alert': '',
        'Patterns': Patterns.objects.all(),
    }

    if request.method == 'POST':
        params = {}
        for key, value in request.POST.items():
            print(f'{key}: {value}')
            params[key] = value

        if 'pattern' not in params:
            data['alert'] = 'Вы не выбрали шаблон'
            return render(request, 'schedule/index.html', data)

        request.session['params'] = params
        return redirect('table')

    return render(request, 'schedule/index.html', data)


def table(request):
    data = {
        'alert': '',
        'Periods': Periods.objects.order_by('time_from'),
        'Absenses': Absenses.objects.order_by('name'),
    }
    params = request.session['params']

    fillParams(data, params)
    #  checkContent(data, params)
    if request.method == 'POST':
        data['graphs'] = {}
        for key, value in request.POST.items():
            if 'graph' in key:
                data['graphs'][key] = value
        data['schedule'] = {name: {day: [] for day in data['days']} for name in data['persons']}
        com = connectPerco(PercoConnect)
        for device in data['devices']:
            getGraphFromPerco(com, device.id_internal, params['date-from'], params['date-to'])
            parceXML(data['schedule'])
        checkContent(data, params)

        fillExcel(data, params)

        return redirect('result')

    return render(request, 'schedule/fill_table.html', data)


def result(request):
    data = {
        'alert': '',
    }

    return render(request, 'schedule/result.html', data)


def download(request):
    with open(f'schedule/files/schedule_{date.today().strftime("%d_%m_%Y")}.xlsx', 'rb') as f:
        file = f.read()

    response = HttpResponse(file, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="schedule_{date.today().strftime("%d_%m_%Y")}.xlsx"'
    return response
